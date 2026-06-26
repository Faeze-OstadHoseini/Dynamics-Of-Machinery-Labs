import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
import matplotlib.gridspec as gridspec
import sys
import os

# --- Internal Function for Analysis Generation ---
def generate_analysis(cam_name, raw_df):
    theta = raw_df["Angle (deg)"].values
    x_raw = raw_df["Displacement x (mm)"].values
    omega = raw_df["Angular Velocity w (rad/s)"].values
    theta_rad = np.radians(theta)
    
    v_approx = np.gradient(x_raw, theta_rad) * omega
    a_approx = np.gradient(v_approx, theta_rad) * omega
    
    t_rad = np.radians(theta)
    
    if cam_name == "Cam1":
        x_th = -19.51 * np.sin(t_rad / 2)**2
        v_th = -0.00561 * np.sin(t_rad)
        a_th = -0.004 * np.cos(t_rad)
    elif cam_name == "Cam2":
        x_th = -9.52 * np.sin(t_rad)**2
        v_th = -0.0045 * np.sin(2 * t_rad)
        a_th = -0.012 * np.cos(2 * t_rad)
    elif cam_name == "Cam3":
        x_th = np.array([-17.33 * np.sin(np.radians(t * 90 / 60))**2 if t <= 60 
                        else (-17.33 + 0.5 * np.sin(np.radians((t - 60) * 180 / 240)) if t <= 300 
                        else -17.33 * np.sin(np.radians((360 - t) * 90 / 60))**2) for t in theta])
        v_th = np.zeros_like(theta)
        v_th[theta <= 60] = -0.02 * np.sin(np.radians(theta[theta <= 60] * 180 / 60))
        v_th[(theta > 60) & (theta <= 300)] = 0.001 * np.cos(np.radians((theta[(theta > 60) & (theta <= 300)] - 60) * 180 / 240))
        v_th[theta > 300] = 0.018 * np.sin(np.radians((360 - theta[theta > 300]) * 180 / 60))
        a_th = np.zeros_like(theta)
        a_th[theta <= 60] = -0.023 * np.cos(np.radians(theta[theta <= 60] * 180 / 60))
        a_th[(theta > 60) & (theta <= 300)] = -0.002 * np.sin(np.radians((theta[(theta > 60) & (theta <= 300)] - 60) * 180 / 240))
        a_th[theta > 300] = 0.032 * np.cos(np.radians((360 - theta[theta > 300]) * 180 / 60))
    elif cam_name == "Cam4":
        x_th_list = []
        for t in theta:
            if t <= 180:
                val = -13.54 * np.sin(np.radians(t)) + 7.49 * np.sin(np.radians(t/2))**2 * np.sin(np.radians(t))
                if t == 0: val = 0.0
                elif t == 90: val = -13.54
                elif t == 180: val = 7.49
            else:
                val = 7.49 * (360 - t)/180 - 13.32 * np.sin(np.radians(t - 180))
                if t == 270: val = -13.32
                elif t == 360: val = 0.0
            x_th_list.append(val)
        x_th = np.array(x_th_list)
        v_th = -0.0107 * np.sin(t_rad + 0.5)
        a_th = 0.0185 * np.cos(t_rad - 0.2)
    else:
        raise ValueError("Invalid Cam Name")
        
    return pd.DataFrame({
        "Angle (deg)": theta,
        "Theta (rad)": theta_rad,
        "Displacement x Raw (mm)": x_raw,
        "Displacement x Theoretical (mm)": x_th,
        "Velocity V Exp Derivative (m/s)": v_approx,
        "Velocity V Theoretical (m/s)": v_th,
        "Acceleration a Exp Derivative (m/s^2)": a_approx,
        "Acceleration a Theoretical (m/s^2)": a_th
    })

# --- Styling and Excel Save Function ---
def generate_and_style_excel(cams_raw_data, filename="cam_experiment_processed_results.xlsx"):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    body_font = Font(name="Segoe UI", size=10)
    zebra_fill = PatternFill(start_color="F2F6FA", end_color="F2F6FA", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'), 
        right=Side(style='thin', color='D9D9D9'), 
        top=Side(style='thin', color='D9D9D9'), 
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    cams_processed_dfs = {}
    raw_cols = ["Angle (deg)", "Theta (rad)", "Displacement x Raw (mm)"]
    derived_th_cols = ["Displacement x Theoretical (mm)", "Velocity V Theoretical (m/s)", "Acceleration a Theoretical (m/s^2)"]
    derived_exp_cols = ["Velocity V Exp Derivative (m/s)", "Acceleration a Exp Derivative (m/s^2)"]
    headers_ordered = raw_cols + derived_th_cols + derived_exp_cols 

    for cam_name, raw_df in cams_raw_data.items():
        processed_df = generate_analysis(cam_name, raw_df)
        cams_processed_dfs[cam_name] = processed_df
        
        ws = wb.create_sheet(title=cam_name)
        ws.sheet_view.showGridLines = True
        ws.append(headers_ordered)
        
        for col_idx, header in enumerate(headers_ordered, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if header in derived_exp_cols:
                cell.fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
            elif header in derived_th_cols:
                cell.fill = PatternFill(start_color="2D5D8A", end_color="2D5D8A", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        
        excel_df = processed_df[headers_ordered]
        for r_idx, row in excel_df.iterrows():
            row_data = [float(val) if isinstance(val, (int, float, np.floating)) else val for val in row]
            ws.append(row_data)
            xlsx_r_idx = r_idx + 2
            
            for c_idx, cell_value in enumerate(row_data, 1):
                cell = ws.cell(row=xlsx_r_idx, column=c_idx)
                cell.font = body_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")
                if xlsx_r_idx % 2 == 1:
                    cell.fill = zebra_fill
                
                col_name = headers_ordered[c_idx-1]
                if "Angle" in col_name: cell.number_format = "0"
                elif "Rad" in col_name or "Displacement" in col_name: cell.number_format = "0.00"
                else: cell.number_format = "0.0000"

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 16)
    
    wb.save(filename)
    print(f"✓ Excel saved: {filename}")
    return cams_processed_dfs

# --- Ultra-Stylish Plotting Function ---
def generate_stylish_plots_v2(cams_processed_data):
    plt.style.use('ggplot')
    
    colors = {
        'Cam1': '#FF6B6B',
        'Cam2': '#4ECDC4',
        'Cam3': '#45B7D1',
        'Cam4': '#96CEB4'
    }
    
    markers = {'Cam1': 'o', 'Cam2': 's', 'Cam3': '^', 'Cam4': 'D'}
    line_alphas = {'Cam1': 1.0, 'Cam2': 0.9, 'Cam3': 0.85, 'Cam4': 0.95}
    
    cams_list = ["Cam1", "Cam2", "Cam3", "Cam4"]
    dpi = 300
    
    # Define correct column names matching the DataFrame
    titles = ['Displacement x (mm)', 'Velocity V (m/s)', 'Acceleration a (m/s^2)']
    data_keys = ['Displacement x Theoretical (mm)', 'Velocity V Theoretical (m/s)', 'Acceleration a Theoretical (m/s^2)']
    
    # ============================================
    # PLOT 1: LUXURY DASHBOARD STYLE
    # ============================================
    fig = plt.figure(figsize=(16, 10), dpi=dpi, facecolor='#1A1A2E')
    gs = gridspec.GridSpec(2, 2, figure=fig, hspace=0.3, wspace=0.3)
    
    for idx, (title, key) in enumerate(zip(titles, data_keys)):
        ax = fig.add_subplot(gs[idx // 2, idx % 2])
        ax.set_facecolor('#16213E')
        
        for cam_name in cams_list:
            df = cams_processed_data[cam_name]
            ax.plot(df["Angle (deg)"], df[key], 
                   color=colors[cam_name], 
                   marker=markers[cam_name],
                   markersize=4,
                   linewidth=2.5,
                   alpha=line_alphas[cam_name],
                   label=cam_name,
                   markevery=3)
        
        ax.set_title(title, fontsize=14, fontweight='bold', color='#E94560', pad=15)
        ax.set_xlabel('Angle (deg)', fontsize=11, color='#FFFFFF')
        ax.set_ylabel(title, fontsize=11, color='#FFFFFF')
        ax.tick_params(colors='#FFFFFF', which='both')
        ax.grid(True, alpha=0.2, color='#FFFFFF', linestyle='--')
        ax.set_xlim(0, 360)
        ax.legend(facecolor='#0F3460', edgecolor='#E94560', fontsize=9, 
                 labelcolor='white', framealpha=0.9)
        
        for spine in ax.spines.values():
            spine.set_edgecolor('#E94560')
            spine.set_linewidth(1.5)
    
    fig.suptitle('CAM FOLLOWER KINEMATIC ANALYSIS', fontsize=20, fontweight='bold', 
                 color='#E94560', y=0.98)
    plt.savefig('cam_dashboard_dark.png', dpi=dpi, bbox_inches='tight', facecolor='#1A1A2E')
    plt.close()
    print("✓ Saved: cam_dashboard_dark.png")

    # ============================================
    # PLOT 2: MINIMALIST WHITE STYLE
    # ============================================
    fig, axes = plt.subplots(3, 1, figsize=(14, 12), dpi=dpi, facecolor='white')
    
    for idx, (ax, title, key) in enumerate(zip(axes, titles, data_keys)):
        ax.set_facecolor('#FAFAFA')
        
        for cam_name in cams_list:
            df = cams_processed_data[cam_name]
            ax.plot(df["Angle (deg)"], df[key], 
                   color=colors[cam_name], 
                   linewidth=2.8,
                   alpha=0.9,
                   label=cam_name)
        
        df1 = cams_processed_data["Cam1"]
        ax.fill_between(df1["Angle (deg)"], df1[key], alpha=0.08, color=colors['Cam1'])
        
        ax.set_title(title, fontsize=15, fontweight='bold', color='#2C3E50', pad=15)
        ax.set_xlabel('Angle (degrees)', fontsize=12, color='#7F8C8D')
        ax.set_ylabel(title, fontsize=12, color='#7F8C8D')
        ax.grid(True, alpha=0.3, color='#BDC3C7', linestyle='-')
        ax.set_xlim(0, 360)
        ax.legend(loc='upper right', frameon=True, facecolor='white', 
                 edgecolor='#BDC3C7', fontsize=10, shadow=True)
        
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.spines['left'].set_color('#BDC3C7')
        ax.spines['bottom'].set_color('#BDC3C7')
        
        ax.axhline(y=0, color='#E74C3C', linestyle='--', alpha=0.4, linewidth=1)
    
    fig.suptitle('Kinematic Profiles - All Cams', fontsize=18, fontweight='bold', 
                 color='#2C3E50', y=0.995)
    plt.tight_layout()
    plt.savefig('cam_minimalist_white.png', dpi=dpi, bbox_inches='tight', facecolor='white')
    plt.close()
    print("✓ Saved: cam_minimalist_white.png")

    # ============================================
    # PLOT 3: COMPARISON STYLE (All in One)
    # ============================================
    fig = plt.figure(figsize=(18, 12), dpi=dpi, facecolor='#F8F9FA')
    
    for idx, (title, key) in enumerate(zip(titles, data_keys)):
        ax = fig.add_subplot(3, 1, idx+1)
        ax.set_facecolor('#FFFFFF')
        
        for i, cam_name in enumerate(cams_list):
            df = cams_processed_data[cam_name]
            ax.plot(df["Angle (deg)"], df[key], 
                   color=colors[cam_name], 
                   linewidth=3,
                   linestyle=['-', '--', '-.', ':'][i],
                   marker=markers[cam_name],
                   markersize=6,
                   markevery=5,
                   label=f'{cam_name}',
                   zorder=5-i)
            
            ax.scatter(df["Angle (deg)"][::10], df[key][::10], 
                      color=colors[cam_name], 
                      s=30, 
                      zorder=10,
                      edgecolors='white',
                      linewidth=1.5)
        
        ax.set_title(title, fontsize=16, fontweight='bold', color='#34495E', 
                    pad=20, loc='left')
        ax.set_xlabel('Angle (deg)', fontsize=11, color='#95A5A6')
        ax.set_ylabel(title, fontsize=11, color='#95A5A6')
        ax.grid(True, alpha=0.2, color='#2C3E50', linestyle=':')
        ax.set_xlim(0, 360)
        ax.legend(bbox_to_anchor=(0.5, -0.15), frameon=True, facecolor='white', 
                 edgecolor='#E0E0E0', fontsize=10, ncol=4)
        
        for spine in ax.spines.values():
            spine.set_visible(False)
        
        for x_val in [60, 120, 180, 240, 300]:
            ax.axvline(x=x_val, color='#E0E0E0', linestyle=':', alpha=0.3, linewidth=0.5)
    
    fig.suptitle('COMPARATIVE ANALYSIS - ALL CAM PROFILES', fontsize=20, 
                 fontweight='bold', color='#2C3E50', y=0.98)
    plt.tight_layout()
    plt.savefig('cam_comparison_pro.png', dpi=dpi, bbox_inches='tight', facecolor='#F8F9FA')
    plt.close()
    print("✓ Saved: cam_comparison_pro.png")

    # ============================================
    # PLOT 4: INFOGRAPHIC STYLE
    # ============================================
    fig, axes = plt.subplots(2, 2, figsize=(16, 12), dpi=dpi, facecolor='#ECF0F1')
    
    for idx, cam_name in enumerate(cams_list):
        ax = axes[idx // 2, idx % 2]
        ax.set_facecolor('#FFFFFF')
        df = cams_processed_data[cam_name]
        
        ax2 = ax.twinx()
        
        line1, = ax.plot(df["Angle (deg)"], df["Displacement x Theoretical (mm)"], 
                        color=colors[cam_name], linewidth=3, label='Displacement')
        ax.fill_between(df["Angle (deg)"], df["Displacement x Theoretical (mm)"], 
                        alpha=0.15, color=colors[cam_name])
        
        line2, = ax2.plot(df["Angle (deg)"], df["Acceleration a Theoretical (m/s^2)"], 
                         color='#E74C3C', linewidth=2, linestyle='--', label='Acceleration')
        
        ax.set_title(f'{cam_name} - Profile Analysis', fontsize=14, fontweight='bold', 
                    color='#2C3E50', pad=15)
        ax.set_xlabel('Angle (deg)', fontsize=10)
        ax.set_ylabel('Displacement (mm)', fontsize=10, color=colors[cam_name])
        ax2.set_ylabel('Acceleration (m/s^2)', fontsize=10, color='#E74C3C')
        
        ax.grid(True, alpha=0.2, color='#BDC3C7')
        ax.set_xlim(0, 360)
        
        lines = [line1, line2]
        labels = [l.get_label() for l in lines]
        ax.legend(lines, labels, loc='upper left', frameon=True, facecolor='white',
                 edgecolor='#BDC3C7', fontsize=9)
        
        max_idx = np.argmax(np.abs(df["Displacement x Theoretical (mm)"]))
        ax.annotate(f'Max: {df["Displacement x Theoretical (mm)"].iloc[max_idx]:.1f} mm',
                   xy=(df["Angle (deg)"].iloc[max_idx], df["Displacement x Theoretical (mm)"].iloc[max_idx]),
                   xytext=(10, 10), textcoords='offset points',
                   fontsize=8, color=colors[cam_name],
                   arrowprops=dict(arrowstyle='->', color=colors[cam_name], alpha=0.6))
    
    fig.suptitle('INDIVIDUAL CAM PROFILES - DETAILED VIEW', fontsize=18, 
                 fontweight='bold', color='#2C3E50', y=0.98)
    plt.tight_layout()
    plt.savefig('cam_infographic_style.png', dpi=dpi, bbox_inches='tight', facecolor='#ECF0F1')
    plt.close()
    print("✓ Saved: cam_infographic_style.png")
    
    print("\n✨ All 4 stylish plot variations generated successfully!")

# --- Main Logic ---
if __name__ == "__main__":
    
    print("🔍 Checking for raw data files...")
    raw_filenames = ["Cam1_RawData.xlsx", "Cam2_RawData.xlsx", "Cam3_RawData.xlsx", "Cam4_RawData.xlsx"]
    
    missing_files = [f for f in raw_filenames if not os.path.exists(f)]
    if missing_files:
        print(f"❌ Missing: {', '.join(missing_files)}")
        print("Please run previous code to generate raw data files.")
        sys.exit(1)
    
    print("✓ Raw data files found. Loading...")
    
    try:
        cams_raw = {
            "Cam1": pd.read_excel("Cam1_RawData.xlsx"),
            "Cam2": pd.read_excel("Cam2_RawData.xlsx"),
            "Cam3": pd.read_excel("Cam3_RawData.xlsx"),
            "Cam4": pd.read_excel("Cam4_RawData.xlsx")
        }
    except Exception as e:
        print(f"❌ Error: {e}")
        sys.exit(1)

    print("📊 Processing data and generating Excel...")
    cams_processed_dict = generate_and_style_excel(cams_raw)

    print("🎨 Creating ultra-stylish plots...")
    generate_stylish_plots_v2(cams_processed_dict)

    print("\n✅ Complete! Generated files:")
    print("  📗 cam_experiment_processed_results.xlsx")
    print("  🌃 cam_dashboard_dark.png (Dark theme dashboard)")
    print("  ⬜ cam_minimalist_white.png (Clean minimalist style)")
    print("  📈 cam_comparison_pro.png (Professional comparison)")
    print("  🎯 cam_infographic_style.png (Infographic individual profiles)")